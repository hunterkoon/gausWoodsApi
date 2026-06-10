"""
CNC Cut Plan - Planilha de Custos (XLSX)
Gera planilha Excel com custos detalhados por peça, por espessura,
totais e aba de COTAÇÃO com resumo completo.

Input:  %TEMP%/cnc_planilha_data.txt
Output: caminho definido em output= no arquivo de dados
"""

import os
import sys
import subprocess
import importlib
from datetime import datetime

try:
    _api_path = os.path.join(os.path.dirname(__file__), "DADOS", "api")
    if _api_path not in sys.path:
        sys.path.insert(0, _api_path)
    from gauswoodsquote.pricing import pv_divisor as _pv_divisor, pv_com_desconto, abaixo_custo, calcular_mao_obra
except ImportError:
    def _pv_divisor(cob, margem_pct, imposto_pct=0.0, comissao_pct=0.0):
        if cob <= 0:
            return 0.0
        soma = min(margem_pct + imposto_pct + comissao_pct, 95.0)
        return cob / (1.0 - soma / 100.0)
    def pv_com_desconto(pv_bruto, desconto_pct):
        if pv_bruto <= 0 or desconto_pct <= 0:
            return pv_bruto
        return pv_bruto * (1.0 - desconto_pct / 100.0)
    def abaixo_custo(pv_final, cob, tolerancia=0.005):
        return pv_final < cob - tolerancia
    def calcular_mao_obra(n_pecas, area_total_m2, tempo_medio_peca_min=12.0, valor_hora=45.0):
        if n_pecas <= 0:
            return 0.0
        return round((n_pecas * tempo_medio_peca_min) / 60.0 * valor_hora, 2)

TEMP_DIR  = os.environ.get("TEMP", os.environ.get("TMP", "/tmp"))
DATA_FILE = os.path.join(TEMP_DIR, "cnc_planilha_data.txt")


# ── Dependência openpyxl ──────────────────────────────────────

def ensure_openpyxl():
    try:
        import openpyxl
        return True
    except ImportError:
        try:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            importlib.invalidate_caches()
            import openpyxl  # noqa
            return True
        except Exception:
            return False


# ── Parser ────────────────────────────────────────────────────

def parse_data(path):
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        lines = [l.rstrip("\n") for l in f]
    meta, chapas, pecas = {}, [], []
    for ln in lines:
        ln_s = ln.strip()
        if not ln_s:
            continue
        if ln_s.startswith("chapa,"):
            p = ln_s.split(",")
            # chapa,esp_mm,qty,price,w_mm,h_mm,prod_nome,acab
            chapas.append({
                "esp":       float(p[1]),
                "qty":       int(float(p[2])),
                "price":     float(p[3]),
                "w":         float(p[4]),
                "h":         float(p[5]),
                "prod_nome": p[6] if len(p) > 6 else "Manual",
                "acab":      p[7] if len(p) > 7 else "",
            })
        elif ln_s.startswith("peca,"):
            p = ln_s.split(",")
            # peca,nome,comp_mm,larg_mm,esp_mm,fita_m
            pecas.append({
                "nome":   p[1],
                "comp":   float(p[2]),
                "larg":   float(p[3]),
                "esp":    float(p[4]),
                "fita_m": float(p[5]) if len(p) > 5 else 0.0,
            })
        elif "=" in ln_s:
            k, v = ln_s.split("=", 1)
            key = k.strip()
            val = v.strip()
            # pecas_plano_json may be very long — store as-is
            meta[key] = val
    return meta, chapas, pecas


def _meta_str(meta: dict, key: str, default: str = "") -> str:
    """Retorna valor de meta como string, ou default se ausente/vazio."""
    v = meta.get(key, default)
    return v if v else default


# ── Motor de cálculo ──────────────────────────────────────────

def calcular(meta, chapas, pecas):
    waste       = float(meta.get("waste_pct",      8.0)) / 100.0   # informativo apenas
    price_fita  = float(meta.get("price_fita_m",   0.0))
    price_ferr  = float(meta.get("price_ferragem",  0.0))
    price_cola  = float(meta.get("price_cola",      0.0))
    price_mo    = float(meta.get("price_mao_obra",  0.0))
    price_frete = float(meta.get("price_frete",     0.0))
    fita_total  = float(meta.get("fita_total_m",    0.0))
    fita_rolo   = int(float(meta.get("fita_rolo_m", 50)))
    fita_nome   = meta.get("fita_nome", "Manual")
    cotacao_id       = int(float(meta.get("cotacao_id",  0)))
    custo_aq_geral   = float(meta.get("custo_aquisicao_geral", 0) or 0)
    custo_prod_geral = float(meta.get("custo_produto_geral",   0) or 0)
    # Modelo v10: PV = COB / (1 - margem - imposto - comissao)  (markup divisor)
    margem_pct    = float(meta.get("margem_pct",   0) or 0)
    imposto_pct   = float(meta.get("imposto_pct",  0) or 0)
    comissao_pct  = float(meta.get("comissao_pct", 0) or 0)
    pv_meta       = float(meta.get("preco_venda",       0) or 0)
    pv_final_meta = float(meta.get("preco_venda_final", 0) or 0)
    fita_consumo_meta = float(meta.get("custo_fita_total",     0) or 0)
    fita_aq_meta      = float(meta.get("custo_fita_aquisicao", 0) or 0)
    cob_meta      = float(meta.get("custo_operacional", 0) or 0)

    # Chapas: custo de aquisição = chapas INTEIRAS consumidas pelo plano.
    # A sobra já está paga dentro da chapa inteira — sem multiplicador de desperdício.
    chapas_out = []
    for c in chapas:
        custo_base  = c["qty"] * c["price"]
        area_total  = c["qty"] * (c["w"] * c["h"] / 1e6)
        pm2 = custo_base / area_total if area_total > 0 else 0.0
        chapas_out.append({**c,
            "custo_base":  round(custo_base,  2),
            "custo_total": round(custo_base,  2),
            "area_m2":     round(area_total,  4),
            "price_m2":    round(pm2,         4)})

    pecas_out  = []
    total_mat  = 0.0
    total_area = 0.0
    total_fita_m = 0.0
    for p in pecas:
        area_p = p["comp"] * p["larg"] / 1e6
        pm2 = 0.0
        for c in chapas_out:
            if abs(c["esp"] - p["esp"]) < 0.5:
                pm2 = c["price_m2"]; break
        mat  = round(area_p * pm2, 4)
        fita = round(p["fita_m"] * price_fita, 4)
        total_mat    += mat
        total_area   += area_p
        total_fita_m += p["fita_m"]
        pecas_out.append({**p, "area_m2": round(area_p, 6),
                          "mat": mat, "fita_cost": fita, "subtotal": round(mat + fita, 4)})

    # Rateio físico: ferragem proporcional à ÁREA da peça; cola proporcional à FITA aplicada.
    # Futuramente, quando ferragens por peça estiverem no input, usar alocação direta.
    n_pecas = max(len(pecas_out), 1)
    for p in pecas_out:
        frac_area = (p["area_m2"] / total_area) if total_area > 0 else (1.0 / n_pecas)
        frac_fita = (p["fita_m"] / total_fita_m) if total_fita_m > 0 else frac_area
        p["ferr_rateio"] = round(price_ferr * frac_area, 4)
        p["cola_rateio"] = round(price_cola * frac_fita, 4)
        p["outros"] = round(p["ferr_rateio"] + p["cola_rateio"], 4)
        p["total"]  = round(p["subtotal"] + p["outros"], 4)

    custo_chapas      = sum(c["custo_base"] for c in chapas_out)
    custo_chapas_base = custo_chapas
    # Sobra de chapa (informativo): diferença entre chapas inteiras pagas e material das peças
    sobra_chapas      = max(0.0, custo_chapas - total_mat)

    # Fita: consumo (entra no CMC) vs aquisição por rolo fechado (entra no CA)
    custo_fita    = fita_consumo_meta if fita_consumo_meta > 0 else round(fita_total * price_fita, 2)
    nr_rolos      = int((fita_total / fita_rolo) + 0.9999) if fita_rolo > 0 else 0
    custo_fita_aq = fita_aq_meta if fita_aq_meta > 0 else round(nr_rolos * fita_rolo * price_fita, 2)
    if custo_fita_aq < custo_fita:
        custo_fita_aq = custo_fita

    # Custo de Aquisição = chapas inteiras + fita por rolo + ferr + cola + frete (SEM MO)
    total_aquisicao = round(custo_chapas + custo_fita_aq + price_ferr + price_cola + price_frete, 2)
    total_geral     = total_aquisicao   # alias mantido para compatibilidade

    # Mao de obra automatica: quando nao informada, calcula por n_pecas
    mo_manual = str(meta.get("mao_obra_manual", "false")).lower() == "true"
    if price_mo <= 0 and not mo_manual and len(pecas_out) > 0:
        price_mo = calcular_mao_obra(len(pecas_out), total_area)
        warnings_mo_auto = True
    else:
        warnings_mo_auto = False

    # CMC (material consumido) e COB (custo operacional base)
    aprov = float(meta.get("aproveitamento_pct", 0) or 0)
    k_perda = min(max(100.0 / aprov - 1.0, 0.0), 2.0) if aprov > 1.0 else 0.10
    warnings = []
    if aprov > 0 and aprov < 50:
        warnings.append(f"Aproveitamento muito baixo ({aprov:.1f}%) — k_perda={k_perda:.2f}")
    if warnings_mo_auto:
        warnings.append(f"Mão de obra calculada automaticamente: R$ {price_mo:.2f} ({len(pecas_out)} peças x 12min x R$45/h)")

    cmc_calc = round(total_mat * (1.0 + k_perda) + custo_fita, 2)
    if custo_prod_geral > 0:
        cmc = custo_prod_geral
        if cmc_calc > 0 and abs(cmc - cmc_calc) / cmc_calc > 0.05:
            warnings.append(f"CMC do meta ({cmc:.2f}) diverge >5% do calculado ({cmc_calc:.2f}) — usando meta")
    else:
        cmc = cmc_calc

    cob_calc = round(cmc + price_ferr + price_cola + price_frete + price_mo, 2)
    if cob_meta > 0:
        cob = cob_meta
        if cob_calc > 0 and abs(cob - cob_calc) / cob_calc > 0.05:
            warnings.append(f"COB do meta ({cob:.2f}) diverge >5% do calculado ({cob_calc:.2f}) — usando meta")
    else:
        cob = cob_calc

    pv = pv_meta if pv_meta > 0 else round(_pv_divisor(cob, margem_pct, imposto_pct, comissao_pct), 2)

    desconto_pct   = float(meta.get("desconto_global", 0) or 0)
    total_com_desc = pv_final_meta if pv_final_meta > 0 else round(pv_com_desconto(pv, desconto_pct), 2)
    desconto_valor = round(pv - total_com_desc, 2)
    venda_abaixo_custo = abaixo_custo(total_com_desc, cob)

    return {
        "chapas": chapas_out, "pecas": pecas_out,
        "custo_chapas":      round(custo_chapas, 2),
        "custo_chapas_base": round(custo_chapas_base, 2),
        "custo_desp":        round(sobra_chapas, 2),   # sobra embutida nas chapas (informativo)
        "custo_fita":        custo_fita,
        "custo_fita_aq":     custo_fita_aq,
        "nr_rolos":          nr_rolos,
        "fita_nome":         fita_nome,
        "fita_rolo":         fita_rolo,
        "price_ferr":        price_ferr,
        "price_cola":        price_cola,
        "price_mo":          price_mo,
        "price_frete":       price_frete,
        "total":             total_geral,
        "total_aquisicao":   total_aquisicao,
        "cmc":               cmc,
        "cob":               cob,
        "margem_pct":        margem_pct,
        "imposto_pct":       imposto_pct,
        "comissao_pct":      comissao_pct,
        "preco_venda":       pv,
        "desconto_pct":      desconto_pct,
        "desconto_valor":    desconto_valor,
        "total_com_desc":    total_com_desc,  # PV final = PV - desconto
        "abaixo_custo":      venda_abaixo_custo,
        "waste_pct":         waste * 100,
        "fita_total":        fita_total,
        "cotacao_id":        cotacao_id,
        "custo_aq_geral":    custo_aq_geral,
        "custo_prod_geral":  custo_prod_geral,
        "mo_auto":           warnings_mo_auto,
        "warnings":          warnings,
    }


# ── Helpers de estilo ─────────────────────────────────────────

MARSALA = "6D1A1A"
DARK    = "1A1A2E"
LGRAY   = "F0F0F0"
WHITE   = "FFFFFF"
GREEN   = "1F4E79"
GOLD    = "C9A84C"


def _fill(color):
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor=color)

def _font(bold=False, color="000000", size=10, italic=False):
    from openpyxl.styles import Font
    return Font(bold=bold, color=color, size=size, italic=italic)

def _align(h="center", v="center", wrap=False):
    from openpyxl.styles import Alignment
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border():
    from openpyxl.styles import Border, Side
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _title_row(ws, row, text, n_cols, color=MARSALA):
    from openpyxl.utils import get_column_letter
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _font(bold=True, color=WHITE, size=13)
    c.fill      = _fill(color)
    c.alignment = _align("center")
    ws.row_dimensions[row].height = 26

def _header_row(ws, row, headers, widths, color=DARK):
    from openpyxl.utils import get_column_letter
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = _font(bold=True, color=WHITE, size=9)
        c.fill      = _fill(color)
        c.alignment = _align("center")
        c.border    = _border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 18

def _data_row(ws, row, values, formats=None, even=True):
    fill    = _fill(LGRAY) if even else _fill(WHITE)
    formats = formats or [""] * len(values)
    for col, (val, fmt) in enumerate(zip(values, formats), 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = fill
        c.border    = _border()
        c.alignment = _align("center" if col > 1 else "left")
        if fmt:
            c.number_format = fmt
    ws.row_dimensions[row].height = 15

def _kv_row(ws, row, label, value, fmt="", label_color=DARK, n_cols=5):
    """Linha label + valor para tabela de resumo."""
    from openpyxl.utils import get_column_letter
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols - 1)}{row}")
    lc = ws.cell(row=row, column=1, value=label)
    lc.font      = _font(bold=True, color=WHITE, size=10)
    lc.fill      = _fill(label_color)
    lc.alignment = _align("left")
    vc = ws.cell(row=row, column=n_cols, value=value)
    vc.font      = _font(bold=True, color=WHITE, size=10)
    vc.fill      = _fill(label_color)
    vc.alignment = _align("center")
    if fmt:
        vc.number_format = fmt
    ws.row_dimensions[row].height = 20

def _money_fmt():
    return 'R$ #.##0,00'

def _br(value):
    return f"{value:.2f}".replace('.', ',')


# ── Abas ──────────────────────────────────────────────────────

def _cli_block(ws, meta, start_row, n_cols=6):
    """
    Insere bloco de dados do cliente (2 colunas: label | valor).
    Retorna o próximo row disponível.
    """
    from openpyxl.utils import get_column_letter

    cli_nome  = _meta_str(meta, "cliente_nome")
    cli_doc   = _meta_str(meta, "cliente_documento")
    cli_tel   = _meta_str(meta, "cliente_telefone")
    cli_email = _meta_str(meta, "cliente_email")
    cli_cid   = _meta_str(meta, "cliente_cidade")
    cli_uf    = _meta_str(meta, "cliente_uf")
    proj      = _meta_str(meta, "nome_projeto")
    entrega   = _meta_str(meta, "previsao_entrega")
    obs       = _meta_str(meta, "obs_geral")

    # Se não há dados de cliente, retorna sem adicionar o bloco
    if not cli_nome and not proj:
        return start_row

    row = start_row
    LBLCOL = "1A3B5F"  # azul escuro para label
    VALCOL = "2C3E50"  # cinza escuro para valor

    def kv(lbl, val):
        nonlocal row
        if not val:
            return
        ws.merge_cells(f"A{row}:B{row}")
        lc = ws.cell(row=row, column=1, value=lbl)
        lc.font = _font(bold=True, color="FFFFFF", size=9)
        lc.fill = _fill(LBLCOL)
        lc.alignment = _align("left")
        ws.merge_cells(f"C{row}:{get_column_letter(n_cols)}{row}")
        vc = ws.cell(row=row, column=3, value=val)
        vc.font = _font(size=9, color="000000")
        vc.fill = _fill("EBF5FB")
        vc.alignment = _align("left")
        ws.row_dimensions[row].height = 16
        row += 1

    kv("Cliente",            cli_nome)
    kv("CPF / CNPJ",         cli_doc)
    kv("Telefone",            cli_tel)
    kv("E-mail",              cli_email)
    cidade_uf = (cli_cid + (" — " + cli_uf if cli_uf else "")) if cli_cid else cli_uf
    kv("Cidade / UF",         cidade_uf)
    kv("Projeto",             proj)
    kv("Previsão de Entrega", entrega)
    kv("Observações",         obs)

    # Linha separadora
    row += 1
    return row


def aba_cotacao(wb, meta, calc):
    """Aba de Cotação: resumo executivo completo para apresentação ao cliente."""
    ws = wb.create_sheet("Cotação", 0)   # primeira aba
    ws.sheet_view.showGridLines = False

    cotacao_id = calc.get("cotacao_id", 0)
    titulo = (f"COTAÇÃO  Nº {cotacao_id:04d}" if cotacao_id > 0
              else "COTAÇÃO — PLANO DE CORTE CNC")

    _title_row(ws, 1, titulo, 6, GREEN)

    ws.merge_cells("A2:F2")
    ws["A2"] = "Data: " + datetime.now().strftime("%d/%m/%Y %H:%M")
    ws["A2"].font      = _font(italic=True, size=9, color="888888")
    ws["A2"].alignment = _align("right")

    # ── Bloco de dados do cliente (se disponível) ──
    row = _cli_block(ws, meta, 4, n_cols=6)
    if row == 4:
        row = 4   # sem dados de cliente — começa direto nos materiais

    # ── Materiais (Chapas) ──
    _title_row(ws, row, "MATERIAIS — CHAPAS", 6, DARK); row += 1
    _header_row(ws, row,
                ["Produto", "Espessura (mm)", "Dim. (mm)", "Qtd", "R$/chapa", "Subtotal"],
                [30, 14, 16, 6, 13, 16])
    row += 1

    for i, c in enumerate(calc["chapas"]):
        _data_row(ws, row,
                  [c["prod_nome"] or "Manual",
                   f'{c["esp"]:.0f}',
                   f'{c["w"]:.0f} × {c["h"]:.0f}',
                   c["qty"],
                   c["price"],
                   c["custo_total"]],
                  ["", "", "", "", _money_fmt(), _money_fmt()],
                  even=(i % 2 == 0))
        row += 1

    ws.cell(row=row, column=5, value="Total chapas:").font = _font(bold=True)
    tc = ws.cell(row=row, column=6, value=calc["custo_chapas"])
    tc.number_format = _money_fmt(); tc.font = _font(bold=True)
    row += 2

    # ── Fita de Borda ──
    _title_row(ws, row, "FITA DE BORDA", 6, DARK); row += 1
    _header_row(ws, row,
                ["Produto", "Total (m)", "Tamanho Rolo", "Nr Rolos", "R$/m", "Subtotal"],
                [30, 11, 13, 9, 12, 16])
    row += 1
    _data_row(ws, row,
              [calc["fita_nome"],
               calc["fita_total"],
               f'{calc["fita_rolo"]} m',
               calc["nr_rolos"],
               float(meta.get("price_fita_m", 0)),
               calc["custo_fita"]],
              ["", "", "", "", _money_fmt(), _money_fmt()])
    row += 2

    # ── Outros custos ──
    _title_row(ws, row, "OUTROS CUSTOS", 6, DARK); row += 1
    _header_row(ws, row, ["Item", "Valor"], [30, 16])
    row += 1
    outros_itens = [
        ("Ferragem",         calc["price_ferr"]),
        ("Cola",             calc["price_cola"]),
        ("Transporte/Frete", calc["price_frete"]),
    ]
    for i, (nome, val) in enumerate(outros_itens):
        _data_row(ws, row, [nome, val], ["", _money_fmt()], even=(i % 2 == 0))
        row += 1
    row += 1

    # ── Totais / Formação do Preço de Venda ──
    def _total_row(label, value, color, size=10, height=20):
        nonlocal row
        ws.merge_cells(f"A{row}:E{row}")
        lbl = ws.cell(row=row, column=1, value=label)
        lbl.font      = _font(bold=True, color="FFFFFF", size=size)
        lbl.fill      = _fill(color)
        lbl.alignment = _align("right")
        val = ws.cell(row=row, column=6, value=value)
        val.number_format = _money_fmt()
        val.font      = _font(bold=True, color="FFFFFF", size=size)
        val.fill      = _fill(color)
        val.alignment = _align("center")
        ws.row_dimensions[row].height = height
        row += 1

    # Custos (visão interna resumida)
    _total_row("Custo de Aquisição dos insumos (sem MO)", calc["total_aquisicao"], DARK)
    _total_row("Custo do Material Consumido (CMC)",       calc["cmc"],             "1A3B5F")
    if calc["price_mo"] > 0:
        _total_row("(+) Mão de Obra (custo interno de fabricação)", calc["price_mo"], "8B6914")
    _total_row("Custo Operacional Base (COB)",            calc["cob"],             "4A235A")

    # Margem + impostos + comissão por dentro do preço (markup divisor)
    pct_lbl = (f"(+) Margem {calc['margem_pct']:.1f}%"
               + (f" + Imposto {calc['imposto_pct']:.1f}%" if calc['imposto_pct'] > 0 else "")
               + (f" + Comissão {calc['comissao_pct']:.1f}%" if calc['comissao_pct'] > 0 else "")
               + " (por dentro do preço)")
    _total_row(pct_lbl, round(calc["preco_venda"] - calc["cob"], 2), "1F4E79")
    _total_row("PREÇO DE VENDA", calc["preco_venda"], "1F618D", size=11, height=24)

    # Desconto (só mostra se > 0) — incide sobre o preço de venda
    if calc["desconto_pct"] > 0:
        _total_row(f"(-) Desconto {calc['desconto_pct']:.1f}% sobre o preço de venda",
                   -calc["desconto_valor"], "C0392B")

    # Alerta: preço final abaixo do custo operacional
    if calc.get("abaixo_custo"):
        ws.merge_cells(f"A{row}:F{row}")
        warn = ws.cell(row=row, column=1,
                       value="ATENÇÃO: o preço final está ABAIXO do custo operacional (COB) — projeto com prejuízo.")
        warn.font      = _font(bold=True, color="FFFFFF", size=10)
        warn.fill      = _fill("922B21")
        warn.alignment = _align("center")
        ws.row_dimensions[row].height = 20
        row += 1

    # Total final (verde)
    GREEN_FINAL = "1E8449"
    _total_row("TOTAL FINAL DO PROJETO", calc["total_com_desc"], GREEN_FINAL, size=12, height=28)
    row += 1

    # ── Rodapé informativo ──
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = (
        f"Aproveitamento do plano: {float(meta.get('aproveitamento_pct', 0)):.1f}%  |  "
        f"Sobra de chapa já paga (informativo): R$ {calc['custo_desp']:.2f}  |  "
        f"Fita total: {calc['fita_total']:.1f} m ({calc['nr_rolos']} rolo(s) de {calc['fita_rolo']} m)"
    )
    ws[f"A{row}"].font      = _font(italic=True, size=9, color="555555")
    ws[f"A{row}"].alignment = _align("center")

    # Freeze após cabeçalho e bloco de cliente
    ws.freeze_panes = "A4"


def aba_resumo(wb, meta, calc):
    ws = wb.create_sheet("Resumo de Custos")
    _title_row(ws, 1, "PLANO DE CUSTOS — CORTE CNC", 5)

    ws["A2"] = "Gerado em: " + datetime.now().strftime("%d/%m/%Y %H:%M")
    ws["A2"].font      = _font(italic=True, size=8, color="888888")
    ws["A2"].alignment = _align("left")

    ws.merge_cells("A3:E3")
    ws["A3"] = "Sobra de chapa (informativo): R$ {:.2f}  |  Fita total: {:.1f} m".format(
                calc["custo_desp"], calc["fita_total"])
    ws["A3"].font      = _font(size=9, color="555555")
    ws["A3"].alignment = _align("center")

    row = 5
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "CHAPAS (inteiras consumidas pelo plano de corte)"
    ws[f"A{row}"].font = _font(bold=True, color=WHITE)
    ws[f"A{row}"].fill = _fill(DARK)
    ws[f"A{row}"].alignment = _align("left")
    ws.row_dimensions[row].height = 18
    row += 1

    _header_row(ws, row,
                ["Espessura", "Qtd", "Preço/chapa",
                 "Subtotal", "Total (chapas inteiras)"],
                [13, 8, 15, 18, 22])
    row += 1

    for i, c in enumerate(calc["chapas"]):
        _data_row(ws, row,
                  [f"{c['esp']:.0f} mm", c["qty"], c["price"], c["custo_base"], c["custo_total"]],
                  ["", "", _money_fmt(), _money_fmt(), _money_fmt()],
                  even=(i % 2 == 0))
        row += 1

    c_total = ws.cell(row=row, column=4, value=sum(c["custo_base"] for c in calc["chapas"]))
    c_total.number_format = _money_fmt(); c_total.font = _font(bold=True); c_total.fill = _fill(LGRAY)
    c_waste = ws.cell(row=row, column=5, value=calc["custo_chapas"])
    c_waste.number_format = _money_fmt(); c_waste.font = _font(bold=True); c_waste.fill = _fill(LGRAY)
    row += 2

    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "OUTROS CUSTOS"
    ws[f"A{row}"].font = _font(bold=True, color=WHITE)
    ws[f"A{row}"].fill = _fill(DARK)
    ws[f"A{row}"].alignment = _align("left")
    ws.row_dimensions[row].height = 18
    row += 1

    outros = [
        ("Fita de Borda (consumo)",   f"{calc['fita_total']:.1f} m", calc["custo_fita"]),
        ("Fita de Borda (aquisição)", f"{calc['nr_rolos']} rolo(s) de {calc['fita_rolo']} m", calc["custo_fita_aq"]),
        ("Ferragem",         "—", calc["price_ferr"]),
        ("Cola",             "—", calc["price_cola"]),
        ("Transporte/Frete", "—", calc["price_frete"]),
        ("Sobra de chapa",   "informativo — já paga nas chapas inteiras", calc["custo_desp"]),
    ]
    for i, (nome, det, val) in enumerate(outros):
        bg = LGRAY if i % 2 == 0 else WHITE
        ws.cell(row=row, column=1, value=nome).fill      = _fill(bg)
        ws.cell(row=row, column=1).alignment             = _align("left")
        ws.cell(row=row, column=2, value=det).fill       = _fill(bg)
        ws.cell(row=row, column=2).font                  = _font(italic=True, size=9)
        c = ws.cell(row=row, column=5, value=val)
        c.number_format = _money_fmt(); c.fill = _fill(bg)
        row += 1

    row += 1

    def _linha_total(label, value, color, size=10, height=20):
        nonlocal row
        ws.merge_cells(f"A{row}:D{row}")
        lc = ws[f"A{row}"]
        lc.value = label
        lc.font = _font(bold=True, color=WHITE, size=size)
        lc.fill = _fill(color)
        lc.alignment = _align("right")
        vc = ws.cell(row=row, column=5, value=value)
        vc.number_format = _money_fmt()
        vc.font = _font(bold=True, color=WHITE, size=size)
        vc.fill = _fill(color); vc.alignment = _align("center")
        ws.row_dimensions[row].height = height
        row += 1

    _linha_total("CUSTO DE AQUISIÇÃO (insumos, sem MO)", calc["total_aquisicao"], MARSALA, size=11, height=22)
    _linha_total("Custo do Material Consumido (CMC)",    calc["cmc"], "1A3B5F")
    if calc["price_mo"] > 0:
        _linha_total("(+) Mão de Obra (custo interno)",  calc["price_mo"], "8B6914")
    _linha_total("Custo Operacional Base (COB)",         calc["cob"], "4A235A")

    pct_lbl = (f"(+) Margem {calc['margem_pct']:.1f}%"
               + (f" + Imposto {calc['imposto_pct']:.1f}%" if calc['imposto_pct'] > 0 else "")
               + (f" + Comissão {calc['comissao_pct']:.1f}%" if calc['comissao_pct'] > 0 else "")
               + " (por dentro do preço)")
    _linha_total(pct_lbl, round(calc["preco_venda"] - calc["cob"], 2), "1F4E79")
    _linha_total("PREÇO DE VENDA", calc["preco_venda"], "1F618D", size=11, height=22)

    if calc.get("desconto_pct", 0) > 0:
        _linha_total(f"(-) Desconto {calc['desconto_pct']:.1f}% sobre o preço de venda",
                     -calc["desconto_valor"], "C0392B")

    if calc.get("abaixo_custo"):
        ws.merge_cells(f"A{row}:E{row}")
        wn = ws[f"A{row}"]
        wn.value = "ATENÇÃO: preço final abaixo do COB — projeto com prejuízo."
        wn.font = _font(bold=True, color=WHITE, size=10)
        wn.fill = _fill("922B21")
        wn.alignment = _align("center")
        ws.row_dimensions[row].height = 20
        row += 1

    GREEN_FINAL = "1E8449"
    _linha_total("TOTAL FINAL DO PROJETO", calc["total_com_desc"], GREEN_FINAL, size=12, height=26)
    row += 2

    # ── Comparativo Custo de Aquisição vs Custo do Produto ──
    custo_aq   = calc.get("custo_aq_geral",  0.0) or calc["total_aquisicao"]
    custo_prod = calc.get("custo_prod_geral", 0.0) or calc["cmc"]
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ANÁLISE DE CUSTO — AQUISIÇÃO vs PRODUTO"
    ws[f"A{row}"].font = _font(bold=True, color=WHITE, size=10)
    ws[f"A{row}"].fill = _fill("1A3B5F")
    ws[f"A{row}"].alignment = _align("center")
    ws.row_dimensions[row].height = 20
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Custo de Aquisição (total desembolsado para comprar os insumos)"
    ws[f"A{row}"].font = _font(size=9)
    ws[f"A{row}"].fill = _fill("EBF5FB")
    ws[f"A{row}"].alignment = _align("left")
    aq_cell = ws.cell(row=row, column=5, value=custo_aq)
    aq_cell.number_format = _money_fmt()
    aq_cell.font = _font(bold=True, size=9)
    aq_cell.fill = _fill("EBF5FB")
    aq_cell.alignment = _align("center")
    ws.row_dimensions[row].height = 16
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Custo do Produto (apenas material consumido no produto — base do preço de venda)"
    ws[f"A{row}"].font = _font(size=9)
    ws[f"A{row}"].fill = _fill("EAFAEA")
    ws[f"A{row}"].alignment = _align("left")
    pr_cell = ws.cell(row=row, column=5, value=custo_prod)
    pr_cell.number_format = _money_fmt()
    pr_cell.font = _font(bold=True, size=9)
    pr_cell.fill = _fill("EAFAEA")
    pr_cell.alignment = _align("center")
    ws.row_dimensions[row].height = 16


def aba_pecas(wb, calc):
    ws = wb.create_sheet("Peças Individuais")
    _title_row(ws, 1, "CUSTO INDIVIDUAL POR PEÇA", 9)

    hdrs = ["Nome", "Comp (mm)", "Larg (mm)", "Esp (mm)",
            "Área (m²)", "Material", "Fita", "Outros", "TOTAL PEÇA"]
    wdts = [22, 11, 11, 10, 10, 14, 12, 12, 14]
    _header_row(ws, 2, hdrs, wdts)

    totals = [0.0] * 4
    for i, p in enumerate(calc["pecas"]):
        r = i + 3
        _data_row(ws, r,
                  [p["nome"], p["comp"], p["larg"], p["esp"],
                   p["area_m2"], p["mat"], p["fita_cost"], p["outros"], p["total"]],
                  ["", "", "", "", "0.000000",
                   _money_fmt(), _money_fmt(), _money_fmt(), _money_fmt()],
                  even=(i % 2 == 0))
        totals[0] += p["mat"];  totals[1] += p["fita_cost"]
        totals[2] += p["outros"]; totals[3] += p["total"]

    tr = len(calc["pecas"]) + 3
    ws.merge_cells(f"A{tr}:E{tr}")
    ws[f"A{tr}"] = "TOTAL"; ws[f"A{tr}"].font = _font(bold=True, color=WHITE)
    ws[f"A{tr}"].fill = _fill(DARK); ws[f"A{tr}"].alignment = _align("right")
    for col, val in zip(range(6, 10), totals):
        c = ws.cell(row=tr, column=col, value=round(val, 2))
        c.number_format = _money_fmt()
        c.font = _font(bold=True, color=WHITE)
        c.fill = _fill(DARK); c.alignment = _align("center")
    ws.row_dimensions[tr].height = 18


def aba_chapas_detalhe(wb, calc):
    ws = wb.create_sheet("Chapas")
    _title_row(ws, 1, "CHAPAS UTILIZADAS POR ESPESSURA", 8)

    _header_row(ws, 2,
                ["Produto", "Espessura", "Dim. (mm)", "Acabamento",
                 "Qtd", "R$/chapa", "Subtotal", "Total (chapas inteiras)"],
                [30, 11, 16, 20, 6, 13, 14, 16])

    for i, c in enumerate(calc["chapas"]):
        _data_row(ws, i + 3,
                  [c["prod_nome"] or "Manual",
                   f"{c['esp']:.0f} mm",
                   f"{c['w']:.0f} × {c['h']:.0f}",
                   c.get("acab", ""),
                   c["qty"], c["price"], c["custo_base"], c["custo_total"]],
                  ["", "", "", "", "", _money_fmt(), _money_fmt(), _money_fmt()],
                  even=(i % 2 == 0))


def aba_pecas_plano(wb, meta):
    """
    Aba 'Peças do Projeto' — lista as peças planejadas com L×A×P, m², custo unitário.
    Lê 'pecas_plano_json' da meta (JSON array serializado pelo MaxScript).
    """
    import json as _json
    raw = meta.get("pecas_plano_json", "[]")
    try:
        pecas_list = _json.loads(raw) if raw else []
        if not isinstance(pecas_list, list):
            pecas_list = []
    except Exception:
        pecas_list = []

    if not pecas_list:
        return  # sem dados — não cria a aba

    ws = wb.create_sheet("Peças do Projeto")
    _title_row(ws, 1, "PEÇAS PLANEJADAS — CUSTO DO PRODUTO", 8, "1E5F1A")
    _header_row(ws, 2,
                ["Nome", "Comp (mm)", "Larg (mm)", "Esp (mm)",
                 "Qtd", "Área m²", "Custo Unit.", "Custo Total"],
                [24, 11, 11, 10, 7, 11, 14, 14],
                color="1E5F1A")
    row = 3
    for i, p in enumerate(pecas_list):
        _data_row(ws, row,
                  [p.get("nome", ""),
                   float(p.get("comp_mm", 0)),
                   float(p.get("larg_mm", 0)),
                   float(p.get("esp_mm",  0)),
                   int(p.get("qtd", 1)),
                   float(p.get("area_m2", 0)),
                   float(p.get("custo_unit", 0)),
                   float(p.get("custo_total", 0))],
                  ["", "", "", "", "", "0.0000", _money_fmt(), _money_fmt()],
                  even=(i % 2 == 0))
        row += 1
    # Totais
    total_custo = sum(float(p.get("custo_total", 0)) for p in pecas_list)
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "TOTAL"
    ws[f"A{row}"].font = _font(bold=True, color=WHITE)
    ws[f"A{row}"].fill = _fill("1E5F1A")
    ws[f"A{row}"].alignment = _align("right")
    tc = ws.cell(row=row, column=8, value=total_custo)
    tc.number_format = _money_fmt()
    tc.font = _font(bold=True, color=WHITE)
    tc.fill = _fill("1E5F1A")
    tc.alignment = _align("center")
    ws.row_dimensions[row].height = 18


def gerar_xlsx(meta, calc, output_path):
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)         # remove aba padrão

    aba_cotacao(wb, meta, calc)          # 1ª aba — Cotação (resumo executivo)
    aba_resumo(wb, meta, calc)           # 2ª aba — Resumo de custos (com comparativo)
    aba_pecas_plano(wb, meta)            # 3ª aba — Peças do projeto (se disponível)
    aba_pecas(wb, calc)                  # 4ª aba — Peças individuais (distribuição de custo)
    aba_chapas_detalhe(wb, calc)         # 5ª aba — Chapas detalhadas

    wb.save(output_path)


# ── Entry point ───────────────────────────────────────────────

def gerar_csv_fallback(meta, calc, out):
    """Gera CSV completo com cotação quando openpyxl não está disponível."""
    cotacao_id = calc.get("cotacao_id", 0)

    with open(out, "w", encoding="utf-8-sig") as f:

        # ── Cabeçalho da cotação ──────────────────────────────
        titulo = f"COTAÇÃO Nº {cotacao_id:04d}" if cotacao_id > 0 else "COTAÇÃO — PLANO DE CORTE CNC"
        f.write(f"{titulo}\n")
        f.write(f"Data;{datetime.now().strftime('%d/%m/%Y %H:%M')}\n")
        f.write(f"Aproveitamento do plano;{meta.get('aproveitamento_pct', '0')} %\n")
        f.write(f"Sobra de chapa (informativo);{_br(calc['custo_desp'])}\n")
        f.write("\n")

        # ── Chapas ───────────────────────────────────────────
        f.write("=== CHAPAS (inteiras consumidas pelo plano) ===\n")
        f.write("Produto;Espessura (mm);Dimensão (mm);Acabamento;Qtd;R$/chapa;Subtotal;Total\n")
        for c in calc["chapas"]:
            f.write(
                f"{c['prod_nome'] or 'Manual'};"
                f"{c['esp']:.0f};"
                f"{c['w']:.0f} x {c['h']:.0f};"
                f"{c.get('acab', '')};"
                f"{c['qty']};"
                f"{_br(c['price'])};"
                f"{_br(c['custo_base'])};"
                f"{_br(c['custo_total'])}\n"
            )
        f.write(f"TOTAL CHAPAS;;;;;; ;{_br(calc['custo_chapas'])}\n")
        f.write("\n")

        # ── Fita de borda ─────────────────────────────────────
        f.write("=== FITA DE BORDA ===\n")
        f.write("Produto;Total (m);Tamanho Rolo;Nº Rolos;R$/m;Subtotal\n")
        f.write(
            f"{calc['fita_nome']};"
            f"{_br(calc['fita_total'])};"
            f"{calc['fita_rolo']} m;"
            f"{calc['nr_rolos']};"
            f"{_br(float(meta.get('price_fita_m', 0)))};"
            f"{_br(calc['custo_fita'])}\n"
        )
        f.write("\n")

        # ── Outros custos ─────────────────────────────────────
        f.write("=== OUTROS CUSTOS ===\n")
        f.write("Item;Valor (R$)\n")
        f.write(f"Ferragem;{_br(calc['price_ferr'])}\n")
        f.write(f"Cola;{_br(calc['price_cola'])}\n")
        f.write(f"Mão de Obra;{_br(calc['price_mo'])}\n")
        f.write(f"Transporte / Frete;{_br(calc['price_frete'])}\n")
        f.write("\n")

        # ── Formação do preço de venda ────────────────────────
        f.write("=== FORMAÇÃO DO PREÇO ===\n")
        f.write(f"Custo de Aquisição (insumos, sem MO);{_br(calc['total_aquisicao'])}\n")
        f.write(f"Custo do Material Consumido (CMC);{_br(calc['cmc'])}\n")
        f.write(f"Custo Operacional Base (COB);{_br(calc['cob'])}\n")
        f.write(f"Margem;{_br(calc['margem_pct'])} %\n")
        f.write(f"Imposto;{_br(calc['imposto_pct'])} %\n")
        f.write(f"Comissão;{_br(calc['comissao_pct'])} %\n")
        f.write(f"PREÇO DE VENDA;{_br(calc['preco_venda'])}\n")
        if calc.get("desconto_pct", 0) > 0:
            f.write(f"Desconto {_br(calc['desconto_pct'])} %;-{_br(calc['desconto_valor'])}\n")
        if calc.get("abaixo_custo"):
            f.write("ATENÇÃO;Preço final abaixo do COB — projeto com prejuízo\n")
        f.write(f"TOTAL FINAL DO PROJETO;{_br(calc['total_com_desc'])}\n")
        f.write("\n")

        # ── Peças individuais ─────────────────────────────────
        f.write("=== PEÇAS INDIVIDUAIS ===\n")
        f.write("Nome;Comp(mm);Larg(mm);Esp(mm);Área(m²);Material(R$);Fita(R$);Outros(R$);Total(R$)\n")
        for p in calc["pecas"]:
            f.write(
                f"{p['nome']};{p['comp']};{p['larg']};{p['esp']};"
                f"{str(round(p['area_m2'], 4)).replace('.', ',')};"
                f"{_br(p['mat'])};{_br(p['fita_cost'])};{_br(p['outros'])};{_br(p['total'])}\n"
            )
        f.write(
            f"TOTAL;;;; ;"
            f"{_br(sum(p['mat'] for p in calc['pecas']))};"
            f"{_br(sum(p['fita_cost'] for p in calc['pecas']))};"
            f"{_br(sum(p['outros'] for p in calc['pecas']))};"
            f"{_br(sum(p['total'] for p in calc['pecas']))}\n"
        )


def main():
    meta, chapas, pecas = parse_data(DATA_FILE)
    calc = calcular(meta, chapas, pecas)
    out  = meta.get("output", "cnc_custos.xlsx")

    if not ensure_openpyxl():
        out_csv = out.replace(".xlsx", ".csv")
        gerar_csv_fallback(meta, calc, out_csv)
        return

    gerar_xlsx(meta, calc, out)


main()
